#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Comentario tecnico: shebang para ejecutar el script con el interprete python3.
# Comentario tecnico: declara codificacion UTF-8 para permitir literales unicode en el archivo.
"""
Motor FormatearIva.
Procesa reporte Amazon (.txt) y base IVA (.csv/.xlsx) para actualizar la base
in-place, generar previsualizacion CSV y un reporte detallado.
"""
# Comentario tecnico: el docstring resume el flujo de procesamiento y efectos en archivos.
# Comentario tecnico: argparse gestiona el parsing de argumentos CLI del motor.
import argparse
# Comentario tecnico: csv aporta lectura y escritura con delimitadores para TXT/CSV.
import csv
# Comentario tecnico: os provee operaciones de filesystem y resolucion de rutas.
import os
import subprocess
import tempfile
import re
import uuid
import zipfile
# Comentario tecnico: sys permite manipular sys.path y escribir errores en stderr.
import sys
# Comentario tecnico: OrderedDict preserva orden de insercion en los mapas de ASIN.
from collections import OrderedDict
# Comentario tecnico: datetime genera marcas de tiempo para el reporte final.
from datetime import datetime
import xml.etree.ElementTree as ET

# Comentario tecnico: nombre por defecto de hoja cuando la base es un XLSX.
BASE_SHEET_NAME = "IVA's Base de Datos"
_LAST_EXCEL_LOG = None
_LAST_EXCEL_VERIFY_LOG = None


# Comentario tecnico: agrega ruta de dependencias locales para que openpyxl sea importable.
def _add_vendor():
    # Comentario tecnico: calcula el directorio absoluto del archivo actual.
    here = os.path.dirname(os.path.abspath(__file__))
    # Comentario tecnico: construye la ruta a la carpeta vendor junto al script.
    vendor = os.path.join(here, 'vendor')
    # Comentario tecnico: inserta vendor al inicio de sys.path solo si existe y no esta.
    if os.path.isdir(vendor) and vendor not in sys.path:
        # Comentario tecnico: prioriza dependencias locales sobre las del entorno global.
        sys.path.insert(0, vendor)


# Comentario tecnico: encapsula la carga de openpyxl desde vendor o entorno.
def _load_openpyxl():
    # Comentario tecnico: asegura que vendor este en sys.path antes de importar.
    _add_vendor()
    try:
        # Comentario tecnico: import dinamico para evitar dependencia dura si no se usa XLSX.
        import openpyxl  # type: ignore
    except Exception as exc:
        # Comentario tecnico: convierte fallas de import en error controlado del motor.
        raise RuntimeError('No se pudo cargar openpyxl desde la carpeta vendor.') from exc
    # Comentario tecnico: retorna el modulo para uso por referencia.
    return openpyxl


# Comentario tecnico: detecta delimitador probable usando conteo en la primera linea.
def _detect_delimiter(sample_line):
    # Comentario tecnico: lista de delimitadores soportados por el motor.
    candidates = ['\t', ';', ',', '|']
    # Comentario tecnico: cuenta ocurrencias por delimitador en la linea de ejemplo.
    counts = {d: sample_line.count(d) for d in candidates}
    # Comentario tecnico: selecciona el delimitador mas frecuente.
    best = max(counts, key=counts.get)
    # Comentario tecnico: si no hay delimitadores, fallback a coma para CSV.
    if counts[best] == 0:
        return ','
    # Comentario tecnico: retorna delimitador detectado.
    return best


# Comentario tecnico: normaliza encabezados para matching estable sin depender de formato original.
def _normalize_header(name):
    # Comentario tecnico: convierte a string, recorta espacios, baja a minusculas y unifica separadores.
    return str(name or '').strip().lower().replace(' ', '-').replace('_', '-')


# Comentario tecnico: evalua si el estado de la orden indica cancelacion.
def _is_cancelled(status):
    # Comentario tecnico: normaliza texto para comparacion insensible a mayusculas.
    s = (status or '').strip().lower()
    # Comentario tecnico: usa substring para capturar variantes de cancelacion.
    return 'cancel' in s


# Comentario tecnico: determina si un valor de impuesto representa IVA positivo.
def _has_tax(value):
    # Comentario tecnico: obtiene string limpio del campo item-tax.
    s = (value or '').strip()
    # Comentario tecnico: campo vacio implica ausencia de impuesto.
    if s == '':
        return False
    try:
        # Comentario tecnico: elimina separadores de miles y parsea a float.
        num = float(s.replace(',', ''))
        # Comentario tecnico: considera impuesto presente si el valor es mayor a cero.
        return num > 0
    except ValueError:
        # Comentario tecnico: si no es numerico pero no esta vacio, se considera presente.
        return True


# Comentario tecnico: normaliza el campo IVA a valores canonicos SI/NO.
def _normalize_iva(value):
    # Comentario tecnico: convierte a mayusculas para comparar variantes.
    s = (value or '').strip().upper()
    # Comentario tecnico: agrupa variantes afirmativas en SI.
    if s in ('SI', 'SÍ', 'YES', 'Y', '1', 'TRUE'):
        return 'SI'
    # Comentario tecnico: agrupa variantes negativas en NO.
    if s in ('NO', 'N', '0', 'FALSE'):
        return 'NO'
    # Comentario tecnico: si no coincide con variantes, devuelve el literal normalizado.
    return s


# Comentario tecnico: lee la primera linea de un archivo de texto con BOM opcional.
def _read_header_line(path):
    # Comentario tecnico: utf-8-sig elimina BOM y errors replace evita fallas por bytes invalidos.
    with open(path, 'r', encoding='utf-8-sig', errors='replace') as f:
        # Comentario tecnico: se lee solo una linea para inferir delimitador.
        line = f.readline()
    # Comentario tecnico: devuelve la linea cruda para parsing posterior.
    return line


# Comentario tecnico: carga la base en CSV y consolida IVA por ASIN preservando orden.
def _load_base_csv(base_path):
    # Comentario tecnico: lee encabezado para inferir delimitador y presencia de columnas.
    header_line = _read_header_line(base_path)
    # Comentario tecnico: si no hay encabezado, el archivo base es invalido.
    if not header_line:
        raise ValueError('El CSV base esta vacio.')
    # Comentario tecnico: detecta delimitador usando la primera linea.
    delimiter = _detect_delimiter(header_line)
    # Comentario tecnico: detecta delimitador final para preservarlo en reescritura.
    trailing_delim = header_line.rstrip('\r\n').endswith(delimiter)

    # Comentario tecnico: parsea encabezados crudos usando csv.reader.
    header_fields = next(csv.reader([header_line], delimiter=delimiter))
    # Comentario tecnico: crea mapa nombre->indice usando normalizacion estable.
    header_map = {_normalize_header(h): i for i, h in enumerate(header_fields) if str(h).strip() != ''}

    # Comentario tecnico: valida columnas requeridas para el motor.
    if 'asin' not in header_map:
        raise ValueError('No se encontro la columna ASIN en el CSV base.')
    if 'iva' not in header_map:
        raise ValueError('No se encontro la columna IVA en el CSV base.')

    # Comentario tecnico: estructura base en memoria para consolidacion.
    base_map = OrderedDict()
    # Comentario tecnico: lista de duplicados para reporte y estadisticas.
    base_duplicates = []
    # Comentario tecnico: contador de filas leidas (sin encabezado).
    total_rows = 0

    # Comentario tecnico: itera todas las filas del CSV base.
    with open(base_path, 'r', encoding='utf-8-sig', errors='replace', newline='') as f:
        # Comentario tecnico: usa el delimitador detectado para lectura correcta.
        reader = csv.reader(f, delimiter=delimiter)
        try:
            # Comentario tecnico: descarta encabezado ya procesado.
            next(reader)
        except StopIteration:
            # Comentario tecnico: si no hay filas, retorna estructuras vacias y metadata.
            return base_map, header_fields, header_map, delimiter, trailing_delim, total_rows, base_duplicates, header_line.rstrip('\r\n')
        # Comentario tecnico: procesa cada fila y consolida duplicados.
        for row in reader:
            # Comentario tecnico: omite filas vacias.
            if not row:
                continue
            # Comentario tecnico: incrementa contador de filas procesadas.
            total_rows += 1
            # Comentario tecnico: completa filas cortas para evitar accesos fuera de rango.
            if len(row) < len(header_fields):
                row += [''] * (len(header_fields) - len(row))
            # Comentario tecnico: extrae ASIN y IVA desde columnas mapeadas.
            asin = row[header_map['asin']].strip()
            iva = row[header_map['iva']].strip()
            # Comentario tecnico: descarta filas sin ASIN.
            if not asin:
                continue
            # Comentario tecnico: normaliza ASIN a mayusculas para clave canonica.
            asin_norm = asin.upper()
            # Comentario tecnico: normaliza IVA con el mismo criterio de salida.
            iva_norm = _normalize_iva(iva)
            # Comentario tecnico: si el ASIN ya existe, consolida priorizando SI.
            if asin_norm in base_map:
                base_duplicates.append(asin_norm)
                if base_map[asin_norm]['IVA'] != 'SI' and iva_norm == 'SI':
                    base_map[asin_norm]['IVA'] = 'SI'
                continue
            # Comentario tecnico: inserta el registro consolidado en el mapa base.
            base_map[asin_norm] = {'ASIN': asin_norm, 'IVA': iva_norm}

    # Comentario tecnico: retorna mapa base y metadata necesaria para reescritura.
    return base_map, header_fields, header_map, delimiter, trailing_delim, total_rows, base_duplicates, header_line.rstrip('\r\n')


# Comentario tecnico: carga la base desde XLSX y consolida IVA por ASIN.
def _load_base_xlsx(base_path, sheet_name):
    # Comentario tecnico: obtiene el modulo openpyxl para manejar XLSX.
    openpyxl = _load_openpyxl()
    # Comentario tecnico: carga el workbook completo desde disco.
    wb = openpyxl.load_workbook(base_path)
    # Comentario tecnico: aplica nombre de hoja por defecto si no se especifica.
    if sheet_name is None:
        sheet_name = BASE_SHEET_NAME
    # Comentario tecnico: valida que la hoja exista y expone alternativas en el error.
    if sheet_name not in wb.sheetnames:
        names = '|'.join(wb.sheetnames)
        raise ValueError('HOJA_NO_ENCONTRADA|' + names)

    # Comentario tecnico: selecciona la hoja para lectura y reescritura.
    ws = wb[sheet_name]
    # Comentario tecnico: extrae encabezados desde la primera fila.
    header_fields = []
    for col in range(1, ws.max_column + 1):
        header_fields.append(ws.cell(row=1, column=col).value)
    # Comentario tecnico: mapa encabezado normalizado -> indice 0-based.
    header_map = {_normalize_header(h): i for i, h in enumerate(header_fields) if str(h or '').strip() != ''}

    # Comentario tecnico: valida columnas requeridas para el motor.
    if 'asin' not in header_map:
        raise ValueError('No se encontro la columna ASIN en la hoja base.')
    if 'iva' not in header_map:
        raise ValueError('No se encontro la columna IVA en la hoja base.')

    # Comentario tecnico: estructura base en memoria para consolidacion.
    base_map = OrderedDict()
    # Comentario tecnico: lista de ASIN duplicados para reporte.
    base_duplicates = []
    # Comentario tecnico: contador de filas con ASIN valido.
    total_rows = 0
    # Comentario tecnico: indices 1-based para acceder a celdas de Excel.
    asin_col = header_map['asin'] + 1
    iva_col = header_map['iva'] + 1

    # Comentario tecnico: recorre filas de datos ignorando el encabezado.
    for row_idx in range(2, ws.max_row + 1):
        # Comentario tecnico: obtiene valores de ASIN e IVA como vienen en la hoja.
        asin_val = ws.cell(row=row_idx, column=asin_col).value
        iva_val = ws.cell(row=row_idx, column=iva_col).value
        # Comentario tecnico: normaliza a string y recorta espacios.
        asin = str(asin_val or '').strip()
        iva = str(iva_val or '').strip()
        # Comentario tecnico: omite filas sin ASIN.
        if not asin:
            continue
        # Comentario tecnico: incrementa contador de filas procesadas.
        total_rows += 1
        # Comentario tecnico: normaliza ASIN a mayusculas para clave canonica.
        asin_norm = asin.upper()
        # Comentario tecnico: normaliza IVA a SI/NO u otros literales.
        iva_norm = _normalize_iva(iva)
        # Comentario tecnico: consolida duplicados priorizando SI.
        if asin_norm in base_map:
            base_duplicates.append(asin_norm)
            if base_map[asin_norm]['IVA'] != 'SI' and iva_norm == 'SI':
                base_map[asin_norm]['IVA'] = 'SI'
            continue
        # Comentario tecnico: inserta nuevo ASIN en el mapa base.
        base_map[asin_norm] = {'ASIN': asin_norm, 'IVA': iva_norm}

    # Comentario tecnico: retorna libro, hoja, metadata y conteos para el flujo principal.
    return wb, ws, sheet_name, base_map, header_fields, header_map, total_rows, base_duplicates


# Comentario tecnico: carga filas del reporte Amazon normalizadas a un dict minimal.
def _load_reporte_rows(reporte_path):
    # Comentario tecnico: lee encabezado para detectar delimitador.
    header_line = _read_header_line(reporte_path)
    # Comentario tecnico: falla si el reporte no contiene encabezados.
    if not header_line:
        raise ValueError('El reporte esta vacio.')
    # Comentario tecnico: detecta delimitador con heuristica de conteo.
    delimiter = _detect_delimiter(header_line)

    # Comentario tecnico: lee el archivo completo con BOM tolerante.
    with open(reporte_path, 'r', encoding='utf-8-sig', errors='replace', newline='') as f:
        # Comentario tecnico: lector CSV con delimitador detectado.
        reader = csv.reader(f, delimiter=delimiter)
        try:
            # Comentario tecnico: obtiene encabezados crudos de la primera fila.
            raw_headers = next(reader)
        except StopIteration:
            # Comentario tecnico: reporte invalido si no hay encabezados.
            raise ValueError('El reporte no tiene encabezados.')

        # Comentario tecnico: normaliza encabezados para acceso estable.
        normalized_headers = [_normalize_header(h) for h in raw_headers]
        # Comentario tecnico: mapea encabezado normalizado a indice.
        header_map = {h: i for i, h in enumerate(normalized_headers)}

        # Comentario tecnico: columnas requeridas por el motor para deducir IVA.
        required = ['asin', 'item-tax', 'order-status']
        # Comentario tecnico: valida columnas faltantes y reporta error detallado.
        missing = [r for r in required if r not in header_map]
        if missing:
            raise ValueError('Faltan columnas en el reporte: ' + ', '.join(missing))

        # Comentario tecnico: itera filas de datos y devuelve solo campos relevantes.
        for row in reader:
            # Comentario tecnico: omite filas vacias.
            if not row:
                continue
            # Comentario tecnico: rellena filas incompletas para proteger indices.
            if len(row) < len(raw_headers):
                row += [''] * (len(raw_headers) - len(row))
            # Comentario tecnico: produce una fila normalizada para el pipeline.
            yield {
                'asin': row[header_map['asin']].strip(),
                'item-tax': row[header_map['item-tax']].strip(),
                'order-status': row[header_map['order-status']].strip(),
            }


# Comentario tecnico: genera un archivo properties plano con claves y valores.
def _write_properties(path, data):
    # Comentario tecnico: usa utf-8 para compatibilidad con caracteres de resumen.
    with open(path, 'w', encoding='utf-8') as f:
        # Comentario tecnico: escribe cada par clave=valor en una linea.
        for key, value in data.items():
            f.write(f"{key}={value}\n")


# Comentario tecnico: detecta y preserva la declaracion XML original si existe.
def _xml_decl(xml_bytes):
    match = re.match(rb'\s*(<\?xml[^>]*\?>)', xml_bytes)
    return match.group(1) if match else None


# Comentario tecnico: serializa XML intentando mantener la declaracion original.
def _serialize_xml(root, original_bytes):
    xml_body = ET.tostring(root, encoding='utf-8')
    decl = _xml_decl(original_bytes)
    if decl:
        return decl + b'\n' + xml_body
    return xml_body


# Comentario tecnico: actualiza core.xml para forzar a Excel a reconocer cambios.
def _update_core_xml(xml_bytes, document_id=None):
    try:
        ns = {
            'cp': 'http://schemas.openxmlformats.org/package/2006/metadata/core-properties',
            'dc': 'http://purl.org/dc/elements/1.1/',
            'dcterms': 'http://purl.org/dc/terms/',
            'dcmitype': 'http://purl.org/dc/dcmitype/',
            'xsi': 'http://www.w3.org/2001/XMLSchema-instance',
        }
        for prefix, uri in ns.items():
            ET.register_namespace(prefix, uri)
        root = ET.fromstring(xml_bytes)
        modified_tag = f"{{{ns['dcterms']}}}modified"
        revision_tag = f"{{{ns['cp']}}}revision"
        modified = root.find(modified_tag)
        if modified is None:
            modified = ET.SubElement(root, modified_tag)
        modified.set(f"{{{ns['xsi']}}}type", 'dcterms:W3CDTF')
        modified.text = datetime.utcnow().replace(microsecond=0).isoformat() + 'Z'
        revision = root.find(revision_tag)
        if revision is None:
            revision = ET.SubElement(root, revision_tag)
            revision.text = '1'
        else:
            try:
                revision.text = str(int(revision.text or '0') + 1)
            except Exception:
                revision.text = '1'
        if document_id:
            identifier_tag = f"{{{ns['dc']}}}identifier"
            identifier = root.find(identifier_tag)
            if identifier is None:
                identifier = ET.SubElement(root, identifier_tag)
            identifier.text = document_id
        return _serialize_xml(root, xml_bytes)
    except Exception:
        return xml_bytes


# Comentario tecnico: elimina la relacion a calcChain para forzar recalc.
def _remove_calc_chain_rel(xml_bytes):
    try:
        ns = {'r': 'http://schemas.openxmlformats.org/package/2006/relationships'}
        ET.register_namespace('', ns['r'])
        root = ET.fromstring(xml_bytes)
        for rel in list(root):
            rel_type = rel.get('Type', '')
            target = rel.get('Target', '')
            if rel_type.endswith('/calcChain') or target.endswith('calcChain.xml'):
                root.remove(rel)
        return _serialize_xml(root, xml_bytes)
    except Exception:
        return xml_bytes


# Comentario tecnico: fuerza recalculo completo al abrir.
def _force_full_calc(xml_bytes):
    try:
        ns = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
        ET.register_namespace('', ns['main'])
        root = ET.fromstring(xml_bytes)
        calc_tag = f"{{{ns['main']}}}calcPr"
        calc_pr = root.find(calc_tag)
        if calc_pr is None:
            calc_pr = ET.SubElement(root, calc_tag)
        calc_pr.set('fullCalcOnLoad', '1')
        calc_pr.set('calcMode', 'auto')
        return _serialize_xml(root, xml_bytes)
    except Exception:
        return xml_bytes


# Comentario tecnico: desactiva actualizaciones automaticas de links/conexiones.
def _update_workbook_pr(xml_bytes):
    try:
        ns = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
        ET.register_namespace('', ns['main'])
        root = ET.fromstring(xml_bytes)
        pr_tag = f"{{{ns['main']}}}workbookPr"
        pr = root.find(pr_tag)
        if pr is None:
            pr = ET.SubElement(root, pr_tag)
        pr.set('updateLinks', 'never')
        pr.set('refreshAllConnections', '0')
        return _serialize_xml(root, xml_bytes)
    except Exception:
        return xml_bytes


# Comentario tecnico: actualiza/crea docProps/custom.xml con un DocumentId.
def _update_custom_xml(xml_bytes, document_id):
    try:
        ns = {
            'cp': 'http://schemas.openxmlformats.org/officeDocument/2006/custom-properties',
            'vt': 'http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes',
        }
        for prefix, uri in ns.items():
            ET.register_namespace(prefix, uri)
        root = ET.fromstring(xml_bytes)
        prop_tag = f"{{{ns['cp']}}}property"
        name_targets = {'documentid', 'document-id', 'docid'}
        max_pid = 1
        found = False
        for prop in root.findall(prop_tag):
            pid = prop.get('pid')
            if pid and pid.isdigit():
                max_pid = max(max_pid, int(pid))
            name = (prop.get('name') or '').strip().lower()
            if name in name_targets:
                for child in list(prop):
                    prop.remove(child)
                val = ET.SubElement(prop, f"{{{ns['vt']}}}lpwstr")
                val.text = document_id
                found = True
        if not found:
            prop = ET.SubElement(root, prop_tag)
            prop.set('fmtid', '{D5CDD505-2E9C-101B-9397-08002B2CF9AE}')
            prop.set('pid', str(max_pid + 1))
            prop.set('name', 'DocumentId')
            val = ET.SubElement(prop, f"{{{ns['vt']}}}lpwstr")
            val.text = document_id
        return _serialize_xml(root, xml_bytes)
    except Exception:
        return xml_bytes


# Comentario tecnico: crea un custom.xml minimo con DocumentId.
def _create_custom_xml(document_id):
    ns_cp = 'http://schemas.openxmlformats.org/officeDocument/2006/custom-properties'
    ns_vt = 'http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes'
    ET.register_namespace('', ns_cp)
    ET.register_namespace('vt', ns_vt)
    root = ET.Element(f"{{{ns_cp}}}Properties")
    prop = ET.SubElement(root, f"{{{ns_cp}}}property")
    prop.set('fmtid', '{D5CDD505-2E9C-101B-9397-08002B2CF9AE}')
    prop.set('pid', '2')
    prop.set('name', 'DocumentId')
    val = ET.SubElement(prop, f"{{{ns_vt}}}lpwstr")
    val.text = document_id
    return _serialize_xml(root, b'')


# Comentario tecnico: asegura relacion de custom-properties en _rels/.rels.
def _ensure_custom_props_relationship(xml_bytes):
    try:
        ns = {'r': 'http://schemas.openxmlformats.org/package/2006/relationships'}
        ET.register_namespace('', ns['r'])
        root = ET.fromstring(xml_bytes)
        rel_tag = f"{{{ns['r']}}}Relationship"
        max_id = 0
        has_custom = False
        for rel in root.findall(rel_tag):
            rel_id = rel.get('Id', '')
            if rel_id.startswith('rId') and rel_id[3:].isdigit():
                max_id = max(max_id, int(rel_id[3:]))
            rel_type = rel.get('Type', '')
            target = rel.get('Target', '')
            if rel_type.endswith('/custom-properties') or target == 'docProps/custom.xml':
                has_custom = True
        if not has_custom:
            rel = ET.SubElement(root, rel_tag)
            rel.set('Id', f'rId{max_id + 1}')
            rel.set('Type', 'http://schemas.openxmlformats.org/officeDocument/2006/relationships/custom-properties')
            rel.set('Target', 'docProps/custom.xml')
        return _serialize_xml(root, xml_bytes)
    except Exception:
        return xml_bytes


# Comentario tecnico: asegura el content type para custom.xml.
def _ensure_custom_props_content_type(xml_bytes):
    try:
        ns = {'ct': 'http://schemas.openxmlformats.org/package/2006/content-types'}
        ET.register_namespace('', ns['ct'])
        root = ET.fromstring(xml_bytes)
        override_tag = f"{{{ns['ct']}}}Override"
        for override in root.findall(override_tag):
            if override.get('PartName') == '/docProps/custom.xml':
                return _serialize_xml(root, xml_bytes)
        override = ET.SubElement(root, override_tag)
        override.set('PartName', '/docProps/custom.xml')
        override.set('ContentType',
                     'application/vnd.openxmlformats-officedocument.custom-properties+xml')
        return _serialize_xml(root, xml_bytes)
    except Exception:
        return xml_bytes


# Comentario tecnico: desactiva refreshOnLoad/refreshOnOpen en XMLs de conexiones.
def _disable_refresh(xml_bytes):
    if (b'refreshOnLoad' not in xml_bytes and b'refreshOnOpen' not in xml_bytes
            and b'refreshOnSave' not in xml_bytes and b'backgroundRefresh' not in xml_bytes):
        return xml_bytes
    try:
        root = ET.fromstring(xml_bytes)
        changed = False
        for elem in root.iter():
            for attr in ('refreshOnLoad', 'refreshOnOpen', 'refreshOnSave', 'backgroundRefresh'):
                val = elem.get(attr)
                if val is not None and val not in ('0', 'false', 'False'):
                    elem.set(attr, '0')
                    changed = True
        if not changed:
            return xml_bytes
        return _serialize_xml(root, xml_bytes)
    except Exception:
        return xml_bytes


# Comentario tecnico: reescribe el XLSX para limpiar caches y metadatos.
def _sanitize_xlsx_file(path):
    temp_path = path + '.san'
    document_id = str(uuid.uuid4()).upper()
    custom_found = False
    try:
        with zipfile.ZipFile(path, 'r') as zin:
            with zipfile.ZipFile(temp_path, 'w') as zout:
                for item in zin.infolist():
                    name = item.filename
                    if name == 'xl/calcChain.xml':
                        continue
                    data = zin.read(name)
                    if name == 'docProps/core.xml':
                        data = _update_core_xml(data, document_id)
                    elif name == 'docProps/custom.xml':
                        data = _update_custom_xml(data, document_id)
                        custom_found = True
                    elif name == '_rels/.rels':
                        data = _ensure_custom_props_relationship(data)
                    elif name == '[Content_Types].xml':
                        data = _ensure_custom_props_content_type(data)
                    elif name == 'xl/_rels/workbook.xml.rels':
                        data = _remove_calc_chain_rel(data)
                    elif name == 'xl/workbook.xml':
                        data = _force_full_calc(data)
                        data = _update_workbook_pr(data)
                    if name.endswith('.xml'):
                        data = _disable_refresh(data)
                    zi = zipfile.ZipInfo(filename=name, date_time=item.date_time)
                    zi.compress_type = item.compress_type
                    zi.comment = item.comment
                    zi.extra = item.extra
                    zi.create_system = item.create_system
                    zi.create_version = item.create_version
                    zi.extract_version = item.extract_version
                    zi.flag_bits = item.flag_bits
                    zi.internal_attr = item.internal_attr
                    zi.external_attr = item.external_attr
                    zout.writestr(zi, data)
                if not custom_found:
                    custom_xml = _create_custom_xml(document_id)
                    zout.writestr('docProps/custom.xml', custom_xml)
        os.replace(temp_path, path)
        return True
    except Exception:
        try:
            if os.path.exists(temp_path):
                os.remove(temp_path)
        except Exception:
            pass
        return False


# Comentario tecnico: escribe la base XLSX usando Excel para sincronizar caches.
def _write_xlsx_with_excel(path, sheet_name, asin_col, iva_col, max_col, records):
    global _LAST_EXCEL_LOG
    if not sys.platform.startswith('win'):
        return False
    flag = os.getenv('IVASINS_EXCEL_WRITE', '1').strip().lower()
    if flag in ('0', 'false', 'no', 'off'):
        return False
    log_flag = os.getenv('IVASINS_EXCEL_LOG', '1').strip().lower()
    log_path = None
    temp_data = None
    temp_script = None
    try:
        temp_dir = os.path.join(tempfile.gettempdir(), 'IvaAsins')
        os.makedirs(temp_dir, exist_ok=True)
        if log_flag not in ('0', 'false', 'no', 'off'):
            stamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
            log_path = os.path.join(temp_dir, f'excel_write_{stamp}.log')
            _LAST_EXCEL_LOG = log_path
        fd, temp_data = tempfile.mkstemp(prefix='ivaasins_', suffix='.csv', dir=temp_dir)
        os.close(fd)
        with open(temp_data, 'w', encoding='utf-8', newline='') as f:
            f.write('ASIN,IVA\r\n')
            for record in records:
                f.write(f"{record['ASIN']},{record['IVA']}\r\n")

        ps_script = r"""
param(
  [string]$path,
  [string]$sheetName,
  [int]$asinCol,
  [int]$ivaCol,
  [int]$maxCol,
  [string]$dataPath,
  [string]$logPath
)
$ErrorActionPreference = "Stop"
$path = [System.IO.Path]::GetFullPath($path)
$logPath = if ($logPath) { [System.IO.Path]::GetFullPath($logPath) } else { $null }
$excel = $null
$wb = $null
function Log([string]$msg) {
    if ($logPath) {
        $ts = (Get-Date).ToString("yyyy-MM-dd HH:mm:ss.fff")
        Add-Content -Path $logPath -Value "$ts $msg"
    }
}
try {
    Log "Start path=$path sheet=$sheetName asinCol=$asinCol ivaCol=$ivaCol maxCol=$maxCol dataPath=$dataPath"
    $excel = New-Object -ComObject Excel.Application
    $excel.DisplayAlerts = $false
    $excel.Visible = $false
    $excel.AskToUpdateLinks = $false
    $excel.EnableEvents = $false
    $excel.AutomationSecurity = 3
    try { $excel.Calculation = -4135 } catch { Log "Calc set failed: $($_.Exception.Message)" }
    Log "Excel version=$($excel.Version)"
    $wb = $excel.Workbooks.Open($path, 0, $false)
    Log "Workbook opened"
    try { $ws = $wb.Worksheets.Item($sheetName) } catch { $ws = $wb.ActiveSheet }
    Log "Worksheet name=$($ws.Name)"
    $used = $ws.UsedRange
    $lastRow = $used.Row + $used.Rows.Count - 1
    if ($lastRow -lt 2) { $lastRow = 1 }
    if ($maxCol -lt 1) { $maxCol = $used.Column + $used.Columns.Count - 1 }
    Log "UsedRange row=$($used.Row) rows=$($used.Rows.Count) col=$($used.Column) cols=$($used.Columns.Count) lastRow=$lastRow maxCol=$maxCol"
    if ($lastRow -ge 2 -and $maxCol -ge 1) {
        $ws.Range($ws.Cells(2, 1), $ws.Cells($lastRow, $maxCol)).ClearContents()
        Log "Cleared range rows=2..$lastRow cols=1..$maxCol"
    }
    $lines = [System.IO.File]::ReadAllLines($dataPath)
    Log "Data lines=$($lines.Length)"
    if ($lines.Length -gt 1) {
        $n = $lines.Length - 1
        $asinArr = New-Object 'object[,]' $n, 1
        $ivaArr = New-Object 'object[,]' $n, 1
        $rowIndex = 0
        for ($i = 1; $i -lt $lines.Length; $i++) {
            $line = $lines[$i]
            if ($line.Length -eq 0) { continue }
            $parts = $line.Split(',', 2)
            $asinArr[$rowIndex, 0] = $parts[0]
            if ($parts.Length -gt 1) { $ivaArr[$rowIndex, 0] = $parts[1] } else { $ivaArr[$rowIndex, 0] = '' }
            $rowIndex++
        }
        if ($rowIndex -gt 0) {
            $ws.Range($ws.Cells(2, $asinCol), $ws.Cells($rowIndex + 1, $asinCol)).Value2 = $asinArr
            $ws.Range($ws.Cells(2, $ivaCol), $ws.Cells($rowIndex + 1, $ivaCol)).Value2 = $ivaArr
            Log "Wrote rows=2..$($rowIndex + 1) to asinCol=$asinCol ivaCol=$ivaCol"
        } else {
            Log "No data rows to write"
        }
    }
    $wb.Save()
    Log "Workbook saved"
    $wb.Close($true)
    Log "Workbook closed"
} catch {
    Log "ERROR: $($_.Exception.Message)"
    throw
} finally {
    if ($wb -ne $null) { [void][System.Runtime.InteropServices.Marshal]::ReleaseComObject($wb) }
    if ($excel -ne $null) { $excel.Quit(); [void][System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel) }
    [GC]::Collect()
    [GC]::WaitForPendingFinalizers()
    Log "COM released"
}
"""
        fd, temp_script = tempfile.mkstemp(prefix='ivaasins_excel_', suffix='.ps1', dir=temp_dir)
        os.close(fd)
        with open(temp_script, 'w', encoding='utf-8', newline='\n') as f:
            f.write(ps_script.strip() + '\n')
        result = subprocess.run(
            ['powershell', '-NoProfile', '-NonInteractive', '-ExecutionPolicy', 'Bypass',
             '-File', temp_script,
             '-path', path,
             '-sheetName', sheet_name or '',
             '-asinCol', str(asin_col),
             '-ivaCol', str(iva_col),
             '-maxCol', str(max_col),
             '-dataPath', temp_data,
             '-logPath', (log_path or '')],
            capture_output=True,
            text=True,
            timeout=300,
        )
        if log_path:
            if result.stdout:
                with open(log_path, 'a', encoding='utf-8') as f:
                    f.write("STDOUT:\n" + result.stdout + "\n")
            if result.stderr:
                with open(log_path, 'a', encoding='utf-8') as f:
                    f.write("STDERR:\n" + result.stderr + "\n")
        return result.returncode == 0
    except Exception:
        return False
    finally:
        if temp_script and os.path.exists(temp_script):
            try:
                os.remove(temp_script)
            except Exception:
                pass
        if temp_data and os.path.exists(temp_data):
            try:
                os.remove(temp_data)
            except Exception:
                pass


# Comentario tecnico: reabre y guarda con Excel para sincronizar caches locales.
def _excel_resave(path):
    if not sys.platform.startswith('win'):
        return False
    flag = os.getenv('IVASINS_EXCEL_RESAVE', '1').strip().lower()
    if flag in ('0', 'false', 'no', 'off'):
        return False
    ps_script = r'''
$ErrorActionPreference = "Stop"
$path = [System.IO.Path]::GetFullPath($args[0])
$excel = $null
$wb = $null
try {
    $excel = New-Object -ComObject Excel.Application
    $excel.DisplayAlerts = $false
    $excel.Visible = $false
    $excel.AskToUpdateLinks = $false
    $excel.EnableEvents = $false
    $excel.AutomationSecurity = 3
    $excel.Calculation = -4135
    $wb = $excel.Workbooks.Open($path, 0, $false)
    $wb.Save()
    $wb.Close($true)
} finally {
    if ($wb -ne $null) { [void][System.Runtime.InteropServices.Marshal]::ReleaseComObject($wb) }
    if ($excel -ne $null) { $excel.Quit(); [void][System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel) }
    [GC]::Collect()
    [GC]::WaitForPendingFinalizers()
}
'''
    try:
        result = subprocess.run(
            ['powershell', '-NoProfile', '-NonInteractive', '-ExecutionPolicy', 'Bypass',
             '-Command', ps_script, path],
            capture_output=True,
            text=True,
            timeout=120,
        )
        return result.returncode == 0
    except Exception:
        return False


# Comentario tecnico: verifica con Excel la ultima fila con datos en columna ASIN.
def _verify_xlsx_with_excel(path, sheet_name, asin_col, expected_rows):
    global _LAST_EXCEL_VERIFY_LOG
    if not sys.platform.startswith('win'):
        return None
    flag = os.getenv('IVASINS_EXCEL_VERIFY', '1').strip().lower()
    if flag in ('0', 'false', 'no', 'off'):
        return None
    log_flag = os.getenv('IVASINS_EXCEL_VERIFY_LOG', '1').strip().lower()
    log_path = None
    temp_script = None
    try:
        temp_dir = os.path.join(tempfile.gettempdir(), 'IvaAsins')
        os.makedirs(temp_dir, exist_ok=True)
        if log_flag not in ('0', 'false', 'no', 'off'):
            stamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
            log_path = os.path.join(temp_dir, f'excel_verify_{stamp}.log')
            _LAST_EXCEL_VERIFY_LOG = log_path
        ps_script = r"""
param(
  [string]$path,
  [string]$sheetName,
  [int]$asinCol,
  [string]$logPath
)
$ErrorActionPreference = "Stop"
$path = [System.IO.Path]::GetFullPath($path)
$logPath = if ($logPath) { [System.IO.Path]::GetFullPath($logPath) } else { $null }
$excel = $null
$wb = $null
function Log([string]$msg) {
    if ($logPath) {
        $ts = (Get-Date).ToString("yyyy-MM-dd HH:mm:ss.fff")
        Add-Content -Path $logPath -Value "$ts $msg"
    }
}
try {
    Log "Verify path=$path sheet=$sheetName asinCol=$asinCol"
    $excel = New-Object -ComObject Excel.Application
    $excel.DisplayAlerts = $false
    $excel.Visible = $false
    $excel.AskToUpdateLinks = $false
    $excel.EnableEvents = $false
    $excel.AutomationSecurity = 3
    $wb = $excel.Workbooks.Open($path, 0, $true)
    try { $ws = $wb.Worksheets.Item($sheetName) } catch { $ws = $wb.ActiveSheet }
    $used = $ws.UsedRange
    $lastUsedRow = $used.Row + $used.Rows.Count - 1
    $lastAsin = $ws.Cells($ws.Rows.Count, $asinCol).End(-4162).Row
    Log "UsedRange row=$($used.Row) rows=$($used.Rows.Count) lastUsedRow=$lastUsedRow lastAsin=$lastAsin"
    Write-Output $lastAsin
    $wb.Close($false)
} finally {
    if ($wb -ne $null) { [void][System.Runtime.InteropServices.Marshal]::ReleaseComObject($wb) }
    if ($excel -ne $null) { $excel.Quit(); [void][System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel) }
    [GC]::Collect()
    [GC]::WaitForPendingFinalizers()
}
"""
        fd, temp_script = tempfile.mkstemp(prefix='ivaasins_verify_', suffix='.ps1', dir=temp_dir)
        os.close(fd)
        with open(temp_script, 'w', encoding='utf-8', newline='\n') as f:
            f.write(ps_script.strip() + '\n')
        result = subprocess.run(
            ['powershell', '-NoProfile', '-NonInteractive', '-ExecutionPolicy', 'Bypass',
             '-File', temp_script,
             '-path', path,
             '-sheetName', sheet_name or '',
             '-asinCol', str(asin_col),
             '-logPath', (log_path or '')],
            capture_output=True,
            text=True,
            timeout=120,
        )
        if log_path:
            if result.stdout:
                with open(log_path, 'a', encoding='utf-8') as f:
                    f.write("STDOUT:\n" + result.stdout + "\n")
            if result.stderr:
                with open(log_path, 'a', encoding='utf-8') as f:
                    f.write("STDERR:\n" + result.stderr + "\n")
        if result.returncode != 0:
            return None
        try:
            return int(result.stdout.strip())
        except Exception:
            return None
    except Exception:
        return None
    finally:
        if temp_script and os.path.exists(temp_script):
            try:
                os.remove(temp_script)
            except Exception:
                pass


# Comentario tecnico: genera un CSV de previsualizacion a partir de nuevos registros.
def _write_preview_csv(path, header_fields, header_map, delimiter, trailing_delim, base_map, start_index):
    # Comentario tecnico: asegura directorio de salida para el archivo preview.
    os.makedirs(os.path.dirname(path), exist_ok=True)
    # Comentario tecnico: normaliza encabezados a string para escritura segura.
    header_fields = [str(h or '') for h in header_fields]
    # Comentario tecnico: construye linea de encabezado con el delimitador original.
    header_line = delimiter.join(header_fields)
    # Comentario tecnico: preserva delimitador final si existia en el CSV base.
    if trailing_delim and (not header_fields or header_fields[-1] != ''):
        header_line += delimiter

    # Comentario tecnico: abre el archivo de salida en modo texto con newline controlado.
    with open(path, 'w', encoding='utf-8', newline='') as f:
        # Comentario tecnico: escribe encabezado en formato CRLF.
        f.write(header_line + '\r\n')
        # Comentario tecnico: itera registros desde el indice de primer agregado.
        for idx, record in enumerate(base_map.values()):
            # Comentario tecnico: descarta registros previos al primer agregado.
            if idx < start_index:
                continue
            # Comentario tecnico: inicializa fila con longitud del encabezado.
            row_out = [''] * len(header_fields)
            # Comentario tecnico: coloca ASIN e IVA en sus columnas respectivas.
            row_out[header_map['asin']] = record['ASIN']
            row_out[header_map['iva']] = record['IVA']
            # Comentario tecnico: serializa fila con delimitador.
            line = delimiter.join(row_out)
            # Comentario tecnico: mantiene delimitador trailing para consistencia.
            if trailing_delim and (not row_out or row_out[-1] != ''):
                line += delimiter
            # Comentario tecnico: escribe fila con terminacion CRLF.
            f.write(line + '\r\n')


# Comentario tecnico: construye un reporte detallado con resumen y listados de cambios.
def _write_report(report_path, base_path, base_type, sheet_name, reporte_path, resumen, added, modified, cancelled_only,
                  base_duplicates, preview_start_index):
    # Comentario tecnico: calcula cantidad de eliminaciones por ASIN duplicado.
    removed_counts = OrderedDict()
    # Comentario tecnico: acumula cantidad de ocurrencias por ASIN duplicado.
    for asin in base_duplicates:
        removed_counts[asin] = removed_counts.get(asin, 0) + 1

    # Comentario tecnico: buffer de lineas para escribir el reporte completo.
    lines = []
    # Comentario tecnico: timestamp para trazabilidad del proceso.
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    # Comentario tecnico: encabezado general del reporte.
    lines.append('REPORTE IVA PROCESS')
    lines.append(f'Fecha/Hora: {now}')
    lines.append('')
    # Comentario tecnico: inicia bloque de resumen global con metadatos del proceso.
    lines.append('RESUMEN GENERAL')
    lines.append(f'Base: {base_path}')
    lines.append(f'Tipo base: {base_type}')
    # Comentario tecnico: agrega nombre de hoja solo cuando la base es XLSX.
    if base_type == 'XLSX':
        lines.append(f'Hoja usada: {sheet_name}')
    lines.append(f'Reporte inventario: {reporte_path}')
    lines.append(f'Total filas en reporte: {resumen["total_reporte"]}')
    lines.append(f'Filas canceladas: {resumen["cancelados_filas"]} (ASIN unicos: {resumen["cancelados_asins"]})')
    lines.append(f'Filas sin ASIN: {resumen["sin_asin_filas"]}')
    lines.append(f'Filas duplicadas en reporte: {resumen["duplicados_filas"]}')
    lines.append(f'ASIN unicos procesados: {resumen["asin_unicos_reporte"]}')
    lines.append(f'Agregados nuevos: {resumen["agregados"]}')
    lines.append(f'Modificados (IVA cambiado): {resumen["modificados"]}')
    lines.append(f'Sin cambios (IVA igual): {resumen["sin_cambios"]}')
    lines.append(f'Duplicados en base consolidados: {resumen["consolidados_base"]}')
    lines.append(f'Eliminados de base (filas): {resumen["eliminados_base"]}')
    lines.append(f'Total base antes: {resumen["base_original"]}')
    lines.append(f'Total base despues: {resumen["base_final"]}')
    lines.append(f'Vista previa inicia en fila (sin encabezado): {preview_start_index + 1}')
    lines.append('')

    # Comentario tecnico: listado de ASIN agregados con su IVA.
    lines.append('PRODUCTOS AGREGADOS (ASIN,IVA)')
    # Comentario tecnico: cabecera de columnas para el listado de agregados.
    lines.append('ASIN,IVA')
    # Comentario tecnico: agrega cada ASIN nuevo al listado.
    for asin, iva in added:
        lines.append(f'{asin},{iva}')
    lines.append('')

    # Comentario tecnico: listado de ASIN modificados y su delta de IVA.
    lines.append('PRODUCTOS MODIFICADOS (ASIN,IVA_ANTERIOR,IVA_NUEVO)')
    # Comentario tecnico: cabecera de columnas para el listado de modificados.
    lines.append('ASIN,IVA_ANTERIOR,IVA_NUEVO')
    # Comentario tecnico: agrega cada ASIN modificado con valores previo y nuevo.
    for asin, iva_old, iva_new in modified:
        lines.append(f'{asin},{iva_old},{iva_new}')
    lines.append('')

    # Comentario tecnico: listado de ASIN omitidos por cancelacion o falta de ASIN.
    lines.append('PRODUCTOS NO PROCESADOS (ASIN,MOTIVO)')
    # Comentario tecnico: cabecera de columnas para el listado de no procesados.
    lines.append('ASIN,MOTIVO')
    # Comentario tecnico: agrega ASIN cancelados que no fueron procesados.
    for asin in sorted(cancelled_only):
        lines.append(f'{asin},CANCELADO')
    # Comentario tecnico: agrega indicador de filas sin ASIN si aplica.
    if int(resumen['sin_asin_filas']) > 0:
        lines.append(f',SIN_ASIN (filas={resumen["sin_asin_filas"]})')
    lines.append('')

    # Comentario tecnico: listado de filas eliminadas por duplicados en base.
    lines.append('PRODUCTOS ELIMINADOS DE LA BASE (ASIN,ELIMINADOS)')
    # Comentario tecnico: cabecera de columnas para el listado de eliminados.
    lines.append('ASIN,ELIMINADOS')
    # Comentario tecnico: agrega conteo de eliminaciones por ASIN duplicado.
    for asin, count in removed_counts.items():
        lines.append(f'{asin},{count}')
    lines.append('')

    # Comentario tecnico: seccion opcional con ASIN duplicados detectados.
    # Comentario tecnico: incluye seccion de duplicados si hubo consolidacion.
    if base_duplicates:
        lines.append('DUPLICADOS EN BASE CONSOLIDADOS (ASIN)')
        # Comentario tecnico: cabecera de columnas para duplicados unicos.
        lines.append('ASIN')
        # Comentario tecnico: lista ASIN duplicados unicos en orden alfabetico.
        for asin in sorted(set(base_duplicates)):
            lines.append(asin)
        lines.append('')

    # Comentario tecnico: asegura directorio destino y escribe el reporte a disco.
    os.makedirs(os.path.dirname(report_path), exist_ok=True)
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))


# Comentario tecnico: orquesta la ejecucion del motor desde la CLI.
def main():
    # Comentario tecnico: define argumentos de entrada y salida del proceso.
    parser = argparse.ArgumentParser(description='Formatear IVA')
    parser.add_argument('--base', required=True, help='CSV/XLSX base con ASIN, IVA')
    parser.add_argument('--reporte', help='Reporte Amazon .txt')
    parser.add_argument('--salida', help='CSV de previsualizacion')
    parser.add_argument('--resumen', help='Archivo resumen (properties)')
    parser.add_argument('--reporte-out', help='Reporte detallado .txt')
    parser.add_argument('--sheet', help='Nombre de hoja para XLSX')
    parser.add_argument('--list-sheets', action='store_true', help='Listar hojas de un XLSX')
    # Comentario tecnico: parsea argumentos CLI a un namespace.
    args = parser.parse_args()

    # Comentario tecnico: resuelve ruta absoluta de la base para validacion.
    base_path = os.path.abspath(args.base)
    # Comentario tecnico: valida que la base exista antes de continuar.
    if not os.path.isfile(base_path):
        raise FileNotFoundError('No existe la base: ' + base_path)

    # Comentario tecnico: modo auxiliar para listar hojas de un XLSX.
    if args.list_sheets:
        # Comentario tecnico: extrae extension para validar tipo de archivo.
        ext = os.path.splitext(base_path)[1].lower()
        # Comentario tecnico: evita listar hojas si la base no es XLSX.
        if ext != '.xlsx':
            raise ValueError('El archivo base no es XLSX.')
        # Comentario tecnico: carga openpyxl para lectura en modo solo lectura.
        openpyxl = _load_openpyxl()
        # Comentario tecnico: abre el workbook sin modificarlo.
        wb = openpyxl.load_workbook(base_path, read_only=True)
        # Comentario tecnico: imprime cada nombre de hoja disponible.
        for name in wb.sheetnames:
            print(name)
        # Comentario tecnico: retorna codigo de salida exitoso sin continuar el flujo.
        return 0

    # Comentario tecnico: valida presencia de argumentos requeridos en modo normal.
    if not args.reporte:
        raise ValueError('Falta --reporte')
    if not args.salida:
        raise ValueError('Falta --salida')

    # Comentario tecnico: resuelve rutas absolutas de entrada y salida.
    reporte_path = os.path.abspath(args.reporte)
    salida_path = os.path.abspath(args.salida)
    # Comentario tecnico: define ruta de resumen y reporte detallado con defaults.
    resumen_path = os.path.abspath(args.resumen) if args.resumen else salida_path + '.resumen'
    reporte_out_path = os.path.abspath(args.reporte_out) if args.reporte_out else None

    # Comentario tecnico: valida existencia del reporte antes de procesarlo.
    if not os.path.isfile(reporte_path):
        raise FileNotFoundError('No existe el reporte: ' + reporte_path)

    # Comentario tecnico: determina el tipo de base por la extension.
    ext = os.path.splitext(base_path)[1].lower()
    base_type = 'XLSX' if ext == '.xlsx' else 'CSV'

    # Comentario tecnico: carga la base y obtiene metadata segun tipo.
    if base_type == 'CSV':
        (base_map, header_fields, header_map, base_delim, trailing_delim,
         base_original_rows, base_duplicates, header_line) = _load_base_csv(base_path)
        # Comentario tecnico: placeholders para el flujo cuando no hay XLSX.
        wb = None
        ws = None
        sheet_name = None
    else:
        wb, ws, sheet_name, base_map, header_fields, header_map, base_original_rows, base_duplicates = _load_base_xlsx(
            base_path, args.sheet)
        # Comentario tecnico: delimitador fijo solo para preview cuando base es XLSX.
        base_delim = ';'
        # Comentario tecnico: no se usa delimitador trailing en XLSX.
        trailing_delim = False

    # Comentario tecnico: mapa del reporte con ASIN unicos y su IVA inferido.
    report_map = OrderedDict()
    # Comentario tecnico: contadores y sets para el resumen final.
    duplicate_rows = 0
    total_rows = 0
    cancelled_rows = 0
    cancelled_asins = set()
    no_asin_rows = 0

    # Comentario tecnico: procesa cada fila del reporte y consolida por ASIN.
    for row in _load_reporte_rows(reporte_path):
        # Comentario tecnico: incrementa contador de filas totales del reporte.
        total_rows += 1
        # Comentario tecnico: obtiene el estado de la orden para filtrar cancelaciones.
        status = row['order-status']
        # Comentario tecnico: omite pedidos cancelados y acumula ASIN cancelados.
        if _is_cancelled(status):
            # Comentario tecnico: incrementa contador de filas canceladas.
            cancelled_rows += 1
            # Comentario tecnico: normaliza ASIN de filas canceladas si existe.
            asin_cancel = row['asin'].strip().upper()
            # Comentario tecnico: agrega el ASIN cancelado solo si no esta vacio.
            if asin_cancel:
                cancelled_asins.add(asin_cancel)
            continue

        # Comentario tecnico: valida ASIN presente en la fila.
        asin = row['asin']
        # Comentario tecnico: contabiliza filas sin ASIN util.
        if not asin:
            no_asin_rows += 1
            continue
        # Comentario tecnico: normaliza ASIN a mayusculas para clave canonica.
        asin_norm = asin.upper()

        # Comentario tecnico: determina IVA segun el monto de item-tax.
        iva_value = 'SI' if _has_tax(row['item-tax']) else 'NO'

        # Comentario tecnico: consolida duplicados en el reporte priorizando SI.
        if asin_norm in report_map:
            # Comentario tecnico: cuenta filas duplicadas en el reporte.
            duplicate_rows += 1
            # Comentario tecnico: si aparece IVA positivo, eleva el valor consolidado.
            if report_map[asin_norm] == 'NO' and iva_value == 'SI':
                report_map[asin_norm] = 'SI'
            continue

        # Comentario tecnico: registra el ASIN con su IVA inferido.
        report_map[asin_norm] = iva_value

    # Comentario tecnico: listas de cambios para el reporte detallado.
    added = []
    modified = []
    # Comentario tecnico: contador de ASIN sin cambios de IVA.
    unchanged = 0
    # Comentario tecnico: indice del primer agregado para preview.
    first_new_index = None

    # Comentario tecnico: aplica el reporte sobre la base en memoria.
    for asin_norm, iva_value in report_map.items():
        if asin_norm in base_map:
            # Comentario tecnico: lee el IVA actual de la base para comparar.
            old_iva = base_map[asin_norm]['IVA']
            # Comentario tecnico: actualiza solo si hay diferencia de IVA.
            if old_iva != iva_value:
                base_map[asin_norm]['IVA'] = iva_value
                modified.append((asin_norm, old_iva, iva_value))
            else:
                unchanged += 1
        else:
            # Comentario tecnico: define el primer indice de agregado.
            if first_new_index is None:
                first_new_index = len(base_map)
            # Comentario tecnico: agrega nuevo ASIN a la base consolidada.
            base_map[asin_norm] = {'ASIN': asin_norm, 'IVA': iva_value}
            added.append((asin_norm, iva_value))

    # Comentario tecnico: si no hubo agregados, el indice inicial es cero.
    if first_new_index is None:
        first_new_index = 0

    # Comentario tecnico: persiste la base con el formato correspondiente.
    if base_type == 'CSV':
        os.makedirs(os.path.dirname(base_path), exist_ok=True)
        with open(base_path, 'w', encoding='utf-8', newline='') as f:
            f.write(header_line + '\r\n')
            # Comentario tecnico: reescribe la base en el orden consolidado.
            for record in base_map.values():
                row_out = [''] * len(header_fields)
                row_out[header_map['asin']] = record['ASIN']
                row_out[header_map['iva']] = record['IVA']
                # Comentario tecnico: serializa la fila con el delimitador detectado.
                line = base_delim.join(row_out)
                # Comentario tecnico: conserva delimitador trailing si corresponde.
                if trailing_delim and (not row_out or row_out[-1] != ''):
                    line += base_delim
                f.write(line + '\r\n')
    else:
        asin_col = header_map['asin'] + 1
        iva_col = header_map['iva'] + 1
        max_col = ws.max_column
        excel_written = _write_xlsx_with_excel(
            base_path,
            sheet_name,
            asin_col,
            iva_col,
            max_col,
            base_map.values(),
        )
        if not excel_written:
            # Comentario tecnico: limpia celdas antiguas y reescribe solo ASIN/IVA.
            max_row = ws.max_row
            # Comentario tecnico: borra celdas de datos preservando encabezados.
            for row_idx in range(2, max_row + 1):
                for col in range(1, max_col + 1):
                    ws.cell(row=row_idx, column=col, value=None)
            # Comentario tecnico: reinicia el indice de fila para escritura.
            row_idx = 2
            # Comentario tecnico: escribe registros consolidados en la hoja.
            for record in base_map.values():
                ws.cell(row=row_idx, column=asin_col, value=record['ASIN'])
                ws.cell(row=row_idx, column=iva_col, value=record['IVA'])
                row_idx += 1
            try:
                wb.active = wb.sheetnames.index(sheet_name)
            except Exception:
                pass
            wb.calculation.fullCalcOnLoad = True
            # Comentario tecnico: guarda en temporal y reemplaza para invalidar cache de Excel.
            temp_path = base_path + '.tmp'
            if os.path.exists(temp_path):
                os.remove(temp_path)
            wb.save(temp_path)
            # Comentario tecnico: limpieza conservadora de caches y conexiones.
            _sanitize_xlsx_file(temp_path)
            os.replace(temp_path, base_path)
            _excel_resave(base_path)
        expected_rows = len(base_map) + 1
        actual_rows = _verify_xlsx_with_excel(base_path, sheet_name, asin_col, expected_rows)
        if actual_rows is not None and actual_rows != expected_rows:
            sys.stderr.write(
                "WARN: Excel muestra %d filas en la columna ASIN, esperado %d. "
                "Revisa consultas/macros. Log: %s\n"
                % (actual_rows, expected_rows, (_LAST_EXCEL_VERIFY_LOG or 'n/a'))
            )

    # Comentario tecnico: genera previsualizacion desde el primer agregado.
    _write_preview_csv(salida_path, header_fields, header_map, base_delim, trailing_delim, base_map, first_new_index)

    # Comentario tecnico: calcula ASIN cancelados que no entraron al mapa final.
    cancelled_only = cancelled_asins - set(report_map.keys())

    # Comentario tecnico: arma el resumen final para integraciones aguas abajo.
    resumen_data = {
        'ok': 'true',
        'total_reporte': str(total_rows),
        'duplicados_filas': str(duplicate_rows),
        'cancelados_filas': str(cancelled_rows),
        'cancelados_asins': str(len(cancelled_only)),
        'sin_asin_filas': str(no_asin_rows),
        'asin_unicos_reporte': str(len(report_map)),
        'agregados': str(len(added)),
        'modificados': str(len(modified)),
        'sin_cambios': str(unchanged),
        'consolidados_base': str(len(set(base_duplicates))),
        'eliminados_base': str(len(base_duplicates)),
        'base_original': str(base_original_rows),
        'base_final': str(len(base_map)),
        'preview_inicio': str(first_new_index),
    }
    # Comentario tecnico: persiste el resumen en disco.
    _write_properties(resumen_path, resumen_data)

    # Comentario tecnico: genera reporte detallado solo si se especifico ruta.
    if reporte_out_path:
        _write_report(
            reporte_out_path,
            base_path,
            base_type,
            sheet_name,
            reporte_path,
            resumen_data,
            added,
            modified,
            cancelled_only,
            base_duplicates,
            first_new_index,
        )

    # Comentario tecnico: imprime estado OK para integracion CLI.
    print('OK')
    # Comentario tecnico: retorna codigo de salida exitoso.
    return 0


if __name__ == '__main__':
    # Comentario tecnico: bloque de entrada cuando se ejecuta como script standalone.
    try:
        # Comentario tecnico: ejecuta main y propaga su codigo de salida al sistema.
        sys.exit(main())
    except Exception as exc:
        # Comentario tecnico: escribe el error en stderr y retorna codigo no cero.
        sys.stderr.write('ERROR: ' + str(exc) + '\n')
        sys.exit(1)
